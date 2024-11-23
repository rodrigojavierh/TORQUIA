import requests
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook, load_workbook
import os

def obtener_torque(url):
    try:
        # Realiza una solicitud GET a la URL
        response = requests.get(url)
        
        # Verifica si la solicitud fue exitosa
        if response.status_code == 200:
            # Analiza el HTML de la respuesta
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Busca todos los elementos de párrafo y verifica si contienen "Torque:"
            torque_element = soup.find('p', string=lambda t: t and "Torque:" in t)
            
            if torque_element:
                # Extrae el valor numérico del torque
                torque_text = torque_element.text.split()[1]  
                torque_value = float(torque_text)
                return torque_value
            else:
                print("No se encontró el valor del torque en la página.")
                return None
        else:
            print("Error al acceder a la página. Código de estado:", response.status_code)
            return None

    except Exception as e:
        print("Error al obtener el torque:", e)
        return None


def guardar_torque_en_excel(torque, archivo="torque_datos.xlsx"):
    """
    Guarda el valor del torque en un archivo Excel.
    
    Parámetros:
        torque (float): Valor del torque a guardar.
        archivo (str): Nombre del archivo Excel (por defecto "torque_datos.xlsx").
    """
    try:
        
        if os.path.exists(archivo):
            wb = load_workbook(archivo)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Tiempo", "Torque (Lb*ft)"])
        
        # Obtener el tiempo actual para registrar cuándo se guardó el torque
        from datetime import datetime
        tiempo_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws.append([tiempo_actual, torque])
        
        wb.save(archivo)
        print(f"Valor de torque guardado en {archivo}")
    
    except Exception as e:
        print("Error al guardar el torque en el archivo Excel:", e)


def verificar_torque_en_rango(torque, minimo=0.01, maximo=1):
    if minimo <= torque <= maximo:
        return True
    else:
        return False

url = "http://192.168.33.175"

while True:
    torque = obtener_torque(url)
    if torque is not None:
        print("Torque Actual:", torque, "Lb*ft")
        
        # Verifica si el torque está dentro, por debajo o por encima del rango
        if torque < 0.01:
            print("El torque está por debajo del rango permitido.")
        elif torque > 1:
            print("El torque está por encima del rango permitido.")
        else:
            print("El perno esta bien ajustado.")
            # Guardar en Excel
            guardar_torque_en_excel(torque)
    else:
        print("No se pudo obtener el torque.")
    
    # Esperar un tiempo antes de la siguiente lectura (1 segundo en este caso)
    time.sleep(0.1)

