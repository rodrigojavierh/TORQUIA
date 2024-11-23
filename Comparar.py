from openpyxl import Workbook, load_workbook
import os

def guardar_torque_en_excel(torque, archivo="torque_datos.xlsx"):
    """
    Guarda el valor del torque en un archivo Excel.
    
    Parámetros:
        torque (float): Valor del torque a guardar.
        archivo (str): Nombre del archivo Excel (por defecto "torque_datos.xlsx").
    """
    try:
        # Verifica si el archivo existe
        if os.path.exists(archivo):
            # Cargar el archivo existente
            wb = load_workbook(archivo)
            ws = wb.active
        else:
            # Crear un nuevo archivo y hoja
            wb = Workbook()
            ws = wb.active
            # Agregar encabezados
            ws.append(["Tiempo", "Torque (Lb*ft)"])
        
        # Obtener el tiempo actual para registrar cuándo se guardó el torque
        from datetime import datetime
        tiempo_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Agregar el valor del torque al archivo
        ws.append([tiempo_actual, torque])
        
        # Guardar el archivo
        wb.save(archivo)
        print(f"Valor de torque guardado en {archivo}")
    
    except Exception as e:
        print("Error al guardar el torque en el archivo Excel:", e)
