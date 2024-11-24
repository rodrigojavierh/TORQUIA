# Programa:
# - CREA IMAGEN DINAMICA A SUPERPONER
# - DETECTA UN MARCADOR ARUCO Y SUPERPONE LA IMAGEN DINAMICA
# - AL PRESIONAR ESPACIO, CAMBIA LOS VALORES DE LA LISTA entregada al inicio Y LA IMÁGEN DINÁNICA
# - los valores cambiaran hasta que se culmine la lista.

#IMPORTAR LIBRERIAS
import cv2
import numpy as np
#import pandas as pd
from recepcion import verificar_torque_en_rango
import requests
from bs4 import BeautifulSoup
import time
import threading
from openpyxl import Workbook, load_workbook
import os


global torque

def obtener_torque(url):
    try:
        # Realiza una solicitud GET a la URL
        response = requests.get(url)
        
        # Verifica si la solicitud fue exitosa
        if response.status_code == 200:
            # Analiza el HTML de la respuesta
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Busca todos los elementos de párrafo y verifica si contienen "Torque:"
            torque_element = soup.find('p', string=lambda t: t and "Torque:" in t)
            
            if torque_element:
                # Extrae el valor numérico del torque
                torque_text = torque_element.text.split()[1]  
                torque_value = float(torque_text)
                return torque_value
            else:
                print("No se encontró el valor del torque")
                return None
        else:
            print("Error al acceder a la página. Código de estado:", response.status_code)
            return None

    except Exception as e:
        print("Error al obtener el torque:", e)
        return None


def guardar_torque_en_excel(torque, archivo="torque_datos.xlsx"):
    """
    Guarda el valor del torque en un archivo Excel.
    
    Parámetros:
        torque (float): Valor del torque a guardar.
        archivo (str): Nombre del archivo Excel (por defecto "torque_datos.xlsx").
    """
    try:
        
        if os.path.exists(archivo):
            wb = load_workbook(archivo)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Tiempo", "Torque (Lb*ft)"])
        
        # Obtener el tiempo actual para registrar cuándo se guardó el torque
        from datetime import datetime
        tiempo_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws.append([tiempo_actual, torque])
        
        wb.save(archivo)
        print(f"Valor de torque guardado en {archivo}")
    
    except Exception as e:
        print("Error al guardar el torque en el archivo Excel:", e)


def verificar_torque_en_rango(torque, minimo=0.01, maximo=1):
    if minimo <= torque <= maximo:
        return True
    else:
        return False

#URL de ip de torquimetro
url = "http://192.168.33.175"

#CREAR DICCIONARIO DE BUSQUEDA DE MARCADORES
parametros = cv2.aruco.DetectorParameters()
diccionario = cv2.aruco.getPredefinedDictionary(cv2.aruco.DICT_5X5_100)

#CONFIG. CÁMARA
cap = cv2.VideoCapture(1)
cap.set(3,1920)
cap.set(4,1080)

#LEER IMAGEN DE SILUETA - Vista 01 - "V01"
im_silueta = cv2.imread("V01_SILU.png")
    # Se podría mejorar cuando se obtiene la info del excel
    # con respecto de la vista en la que se está trabajando


flag_mod_torqueado=0
tornillo_a_ajustar=1 #info que se lee inicialmente del excel
v_estado = [1,0,0,0,0,0,0,0] #2:verde 1:azul 0:rojo #info que se lee inicialmente del excel
# Task principal que incluye toda la lógica
def task_procesar_tornillos():
    # Emulación de datos iniciales
    cant_tornillos = 8  # Cantidad de tornillos en la vista
    listaX = [76, 246, 411, 578, 91, 250, 418, 583]
    listaY = [65, 65, 62, 63, 258, 230, 225, 225]
    v_estado = [0] * cant_tornillos  # Estado inicial de los tornillos
    tornillo_a_ajustar = 1
    flag_mod_torqueado = 0  # Indicador de estado de torqueado
    im_silueta = np.zeros((300, 600, 3), dtype=np.uint8)  # Imagen base (silueta)
    diccionario = cv2.aruco.Dictionary_get(cv2.aruco.DICT_4X4_50)
    parametros = cv2.aruco.DetectorParameters_create()
    cap = cv2.VideoCapture(0)

    while True:
        # Actualización de estados de los tornillos
        if flag_mod_torqueado == 1:
            tornillo_a_ajustar += 1
            v_estado[tornillo_a_ajustar - 2] = 2
            v_estado[tornillo_a_ajustar - 1] = 1
        if flag_mod_torqueado == 2:
            v_estado[tornillo_a_ajustar - 1] = 2

        # Crear imagen negra de tornillos
        alto, ancho, canales = im_silueta.shape
        imagen_negra = np.zeros((alto, ancho, canales), dtype=np.uint8)
        for i in range(0, cant_tornillos):
            if v_estado[i] == 2:  # Verde
                color_circulo = (0, 255, 0)
            elif v_estado[i] == 1:  # Azul
                color_circulo = (255, 0, 0)
            elif v_estado[i] == 0:  # Rojo
                color_circulo = (0, 0, 255)
            else:
                color_circulo = (0, 0, 0)
            cv2.circle(imagen_negra, (listaX[i], listaY[i]), 20, color_circulo, 2)

        # Sumar imágenes con datos de tornillos y silueta
        im_silu_mod = cv2.add(im_silueta, imagen_negra)

        while True:
            # Leer imágenes de cámara y transformar a escala de grises
            ret, frame = cap.read()
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

            # Procesamiento de imagen y detección de marcadores
            esquinas, ids, _ = cv2.aruco.detectMarkers(gray, diccionario, parameters=parametros)

            # Dibujar marcador detectado
            if np.all(ids != None):  # Si existe marcador ArUco
                aruco = cv2.aruco.drawDetectedMarkers(frame, esquinas)

                # Superponer imagen según escalado
                sc_x, sc_y = 15, 8
                c1 = (esquinas[0][0][0][0], esquinas[0][0][0][1])
                c2 = (sc_x * esquinas[0][0][1][0] - (sc_x - 1) * esquinas[0][0][0][0],
                      sc_y * esquinas[0][0][1][1] - (sc_y - 1) * esquinas[0][0][0][1])
                c3 = (sc_x * esquinas[0][0][2][0] - (sc_x - 1) * esquinas[0][0][0][0],
                      sc_y * esquinas[0][0][2][1] - (sc_y - 1) * esquinas[0][0][0][1])
                c4 = (sc_x * esquinas[0][0][3][0] - (sc_x - 1) * esquinas[0][0][0][0],
                      sc_y * esquinas[0][0][3][1] - (sc_y - 1) * esquinas[0][0][0][1])

                # Leer imagen a superponer
                imagen = im_silu_mod
                tamanho = imagen.shape
                puntosRef_aruco = np.array([c1, c2, c3, c4])
                puntos_imagen = np.array([
                    [0, 0],
                    [(tamanho[1] - 1), 0],
                    [(tamanho[1] - 1), (tamanho[0] - 1)],
                    [0, (tamanho[0] - 1)]
                ], dtype=float)

                # Superponer imagen transformada
                h, _ = cv2.findHomography(puntos_imagen, puntosRef_aruco)
                perspectiva = cv2.warpPerspective(imagen, h, (frame.shape[1], frame.shape[0]))
                frame = cv2.addWeighted(src1=frame, alpha=1, src2=perspectiva, beta=1, gamma=0)
                cv2.imshow("Realidad Virtual", frame)
            else:
                cv2.imshow("Realidad Virtual", frame)

            # Verificar torque dentro del rango
            
            bandera = verificar_torque_en_rango(torque)
            if bandera:
                print("Torque dentro del rango:", torque)
                break
            else:
                print("Torque fuera del rango:", torque)

            if cv2.waitKey(1) == 27:  # ESC para salir del bucle interno
                break

        # Finalizar lógica de tornillos
        if len(v_estado) == tornillo_a_ajustar:
            flag_mod_torqueado = 2
        else:
            flag_mod_torqueado = 1

        if cv2.waitKey(1) == 27:  # ESC para salir del bucle principal
            break

    # Liberar cámara y cerrar ventanas
    cap.release()
    cv2.destroyAllWindows()
# Función 1: Monitoreo de torque
def monitoreo_torque(url):
    while True:
        torque = obtener_torque(url)
        if torque is not None:
            print("Torque Actual:", torque, "Lb*ft")
            
            # Verifica si el torque está dentro, por debajo o por encima del rango
            if torque < 0.01:
                print("El torque está por debajo del rango permitido.")
            elif torque > 1:
                print("El torque está por encima del rango permitido.")
            else:
                print("El perno está bien ajustado.")
                # Guardar en Excel
                guardar_torque_en_excel(torque)
        else:
            print("No se pudo obtener el torque.")
        
        # Esperar un tiempo antes de la siguiente lectura (1 segundo en este caso)
        time.sleep(0.1)
        if cv2.waitKey(1) == 27:  # ESC para salir del bucle principal
            break




# Crear el hilo para ejecutar el task
thread_tornillos = threading.Thread(target=task_procesar_tornillos)
thread_torque = threading.Thread(target= monitoreo_torque)
# Iniciar el hilo
thread_tornillos.start()
thread_torque.start()

# Esperar a que el hilo termine (opcional, si el programa es infinito no es necesario)
thread_tornillos.join()
thread_torque.join()
